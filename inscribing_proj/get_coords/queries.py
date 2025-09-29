import re
from openpyxl import load_workbook
from pymongo import MongoClient, GEOSPHERE
import bisect
import requests
import json
from decouple import config

database_id = config("DATABASE_ID")
datacube_api_url = config("BASE_DATACUBE_URL")
crud_url = datacube_api_url+"api/crud"
api_key=config("API_KEY")

headers = {
        "Authorization": f"Api-Key {api_key}",
        "Content-Type": "application/json"
    }
### DATACUBE QUERIES ###
def get_data_datacube(collection_name=config("INDEX_COLLECTION_NAME"),filters={}):
    length =len( list(filters.keys()))
    params =params = {
    "database_id": database_id,
    "collection_name": collection_name,
    "filters": filters
}

    print(f"Get data datacube filters = {filters} length of filters = {length} collection_name {collection_name}")
    if length:
        filters = json.dumps(filters)
    try:
        res = requests.get(url=crud_url,params=params, headers=headers).json()
        if res['success']:
            return res
        else:
            raise Exception
    except Exception as e:
        print(f"Exceptions is {e}")
def sort_data_cube(res):
    print("Sorting lats")
    sorted_lats = sorted([doc['latitude'] for doc in res['data']])
    return sorted_lats
def get_data_from_db_datacube(target_lats ,lat_max,lat_min,lon_max,lon_min, raw_data):
    print("---Inside in get datacuve")
    # print(f"Now target lats = {target_lats}")
    collections = [doc['collection'] for doc in raw_data['data'] if doc['latitude'] in target_lats]
    # print(f"Target collections {collections}")
    
    # Query databases
    raw_results = []
    geo_results = []
    print("---packing results")
    for coll_name in collections:
        # Raw database query
        raw_docs = get_data_datacube(collection_name=coll_name,filters={
            'longitude': {'$gte': lon_min, '$lte': lon_max}
        })

        for doc in raw_docs:
            doc.pop('_id', None)
        raw_results.extend(raw_docs)
        
      
    print("--returning results")
    
    return {
        'raw_coordinates': raw_results
    }
def query_by_four_corners_datacube(top_left, top_right, bottom_left, bottom_right):
    """
    Query coordinates within bounding box defined by four (lat, lon) points.
    Uses binary search on sorted latitude index for efficiency.
    """
    # Extract all points (each as (lat, lon))
    points = [top_left, top_right, bottom_left, bottom_right]
    print(f"----received Points {points}")

    # Calculate bounding box
    lats = [float(p[0] )for p in points]
    lons = [float(p[1]) for p in points]
    lat_min, lat_max = min(lats), max(lats)
    lon_min, lon_max = min(lons), max(lons)
    raw_data = get_data_datacube()
    try:
        if raw_data["success"]:
            sorted_lats = sort_data_cube(raw_data)
            target_lats = bin_search(sorted_lats, top_left,top_right,bottom_left,
                                     bottom_right,lat_max,lat_min)
            results = get_data_from_db_datacube(target_lats,lat_max, lat_min,lon_max, lon_min,raw_data)
            return results
        else:
            print(f"No success!")
            raise Exception
    except Exception as e:
        print(f"Exceptions is {e}")
        # return {"success":False, "error":True, "message": f"There was an error during accessing datacube {e}"}
    
        # results = get_data_from_db(target_lats,lat_max,lat_min,lon_max,lon_min)
        # return results
### END OF DATACUVBE
def store_coordinates(file_path, mongo_uri='mongodb://localhost:27017/'):
    """
    Reads coordinates from Excel, stores in MongoDB with correct (lat, lon) order,
    and creates optimized index collections.
    """
    client = MongoClient(mongo_uri)
    db_raw = client['raw_coordinates_db']
    db_geo = client['geocoordinates_db']
    
    # Create unified index collection
    index_db = client['coordinate_index_db']
    lat_index_coll = index_db['latitude_index']
    lat_index_coll.create_index("latitude")
    
    # Track processed latitudes
    processed_lats = set()
    
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
        for col_idx, cell in enumerate(row[1:], start=1):
            value = cell.value
            if not value or not isinstance(value, str):
                continue
                
            match = re.match(r'\(([^,]+),\s*([^)]+)\)', value.strip())
            if not match:
                continue
                
            # Extract and swap values: first is latitude, second is longitude
            lat_str, lon_str = match.groups()
            try:
                lat_val = float(lat_str.strip())
                lon_val = float(lon_str.strip())
            except ValueError:
                continue
                
            # Create collection name from latitude
            collection_name = f"lat_{lat_val:.6f}"
            
            # Insert into raw database
            raw_coll = db_raw[collection_name]
            raw_coll.insert_one({
                'latitude': lat_val,
                'longitude': lon_val
            })
            
            # Insert into GeoJSON database
            geo_coll = db_geo[collection_name]
            geo_coll.insert_one({
                'location': {
                    'type': 'Point',
                    'coordinates': [lon_val, lat_val]  # GeoJSON: [lon, lat]
                }
            })
            
            # Create geospatial index for new collections
            if lat_val not in processed_lats:
                geo_coll.create_index([('location', GEOSPHERE)])
                # Add to unified index
                lat_index_coll.insert_one({
                    'latitude': lat_val,
                    'collection': collection_name
                })
                processed_lats.add(lat_val)
    
    wb.close()
    client.close()
def bin_search(sorted_lats,top_left, top_right, bottom_left, bottom_right, lat_max, lat_min):
    # Find latitude range using binary search
    left_idx = bisect.bisect_left(sorted_lats, lat_min)
    print(f"left idx(lat_min) = {left_idx}")
    right_idx = bisect.bisect_right(sorted_lats, lat_max)
    print(f"right idx (lat_max))= {right_idx}")

    target_lats = sorted_lats[left_idx:right_idx+1]
    # print(f"target_lats= {target_lats}")
    return target_lats
def get_index( mongo_uri='mongodb://localhost:27017/'):
        client = MongoClient(mongo_uri)
        index_db = client['coordinate_index_db']
         # Get sorted latitude index
        lat_index = index_db['latitude_index']
        print("----Sorting index----")
        sorted_lats = sorted([doc['latitude'] for doc in lat_index.find({})])
        print("----Searching")
        return sorted_lats
def get_data_from_db(target_lats ,lat_max,lat_min,lon_max,lon_min, mongo_uri='mongodb://localhost:27017/'):
    client = MongoClient(mongo_uri)
    db_raw = client['raw_coordinates_db']
    db_geo = client['geocoordinates_db']
    index_db = client['coordinate_index_db']
    print("---getting collection names")
    lat_index = index_db['latitude_index']
    # Get collection names
    collections = [doc['collection'] for doc in 
                  lat_index.find({'latitude': {'$in': target_lats}})]
    print(f"Target collections {collections}")
    
    # Query databases
    raw_results = []
    geo_results = []
    print("---packing results")
    for coll_name in collections:
        # Raw database query
        raw_coll = db_raw[coll_name]
        raw_docs = list(raw_coll.find({
            'longitude': {'$gte': lon_min, '$lte': lon_max}
        }))
        for doc in raw_docs:
            doc.pop('_id', None)
        raw_results.extend(raw_docs)
        
        # GeoJSON database query
        geo_coll = db_geo[coll_name]
        geo_docs = list(geo_coll.find({
            'location': {
                '$geoWithin': {
                    '$geometry': {
                        'type': 'Polygon',
                        'coordinates': [[
                            [lon_min, lat_min],
                            [lon_max, lat_min],
                            [lon_max, lat_max],
                            [lon_min, lat_max],
                            [lon_min, lat_min]
                        ]]
                    }
                }
            }
        }))
        for doc in geo_docs:
            doc.pop('_id', None)
        geo_results.extend(geo_docs)
    
    client.close()
    print("--returning results")
    
    return {
        'raw_coordinates': raw_results,
        'geo_coordinates': geo_results
    }
def query_by_four_corners(top_left, top_right, bottom_left, bottom_right):
    """
    Query coordinates within bounding box defined by four (lat, lon) points.
    Uses binary search on sorted latitude index for efficiency.
    """
    # Extract all points (each as (lat, lon))
    points = [top_left, top_right, bottom_left, bottom_right]
    # print(f"points = {points}")
    print("----received Points")
    # Calculate bounding box
    lats = [float(p[0] )for p in points]
    lons = [float(p[1]) for p in points]
    lat_min, lat_max = min(lats), max(lats)
    lon_min, lon_max = min(lons), max(lons)
    # print(f"type lat_min {type(lat_min)} == {lat_min}")
    sorted_lats = get_index()
    target_lats = bin_search(sorted_lats, top_left,top_right,bottom_left,bottom_right,lat_max,lat_min)
    results = get_data_from_db(target_lats,lat_max,lat_min,lon_max,lon_min)
    return results
    
    
   




# store_coordinates('scaled_coords_200_200.xlsx')