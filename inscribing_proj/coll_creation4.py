import re
from openpyxl import load_workbook
from pymongo import MongoClient, GEOSPHERE
import bisect
import requests
import json
from decouple import config

datacube_url = config("BASE_DATACUBE_URL")
database_id = config("DATABASE_ID")

crud_url=datacube_url+"/api/crud/"
create_url=datacube_url+"/api/create_database/"
add_colls_url = datacube_url+"/api/add_collection/"
api_key=config("API_KEY")

headers = {
        "Authorization": f"Api-Key {api_key}",
        "Content-Type": "application/json"
    }
def store_coordinates(file_path, mongo_uri='mongodb://localhost:27017/'):
    """
    Reads coordinates from Excel, stores in MongoDB with correct (lat, lon) order,
    and creates optimized index collections.
    """
    client = MongoClient(mongo_uri)
    db_raw = client['raw_coordinates_db']
    db_geo = client['geocoordinates_db']
    
    # Create unified index collection
    index_db = client['coordinate_index_db']
    lat_index_coll = index_db['latitude_index']
    lat_index_coll.create_index("latitude")
    
    # Track processed latitudes
    processed_lats = set()
    
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
        for col_idx, cell in enumerate(row[1:], start=1):
            value = cell.value
            if not value or not isinstance(value, str):
                continue
                
            match = re.match(r'\(([^,]+),\s*([^)]+)\)', value.strip())
            if not match:
                continue
                
            # Extract and swap values: first is latitude, second is longitude
            lat_str, lon_str = match.groups()
            try:
                lat_val = float(lat_str.strip())
                lon_val = float(lon_str.strip())
            except ValueError:
                continue
                
            # Create collection name from latitude
            collection_name = f"lat_{lat_val:.6f}"
            
            # Insert into raw database
            raw_coll = db_raw[collection_name]
            raw_coll.insert_one({
                'latitude': lat_val,
                'longitude': lon_val
            })
            
            # Insert into GeoJSON database
            geo_coll = db_geo[collection_name]
            geo_coll.insert_one({
                'location': {
                    'type': 'Point',
                    'coordinates': [lon_val, lat_val]  # GeoJSON: [lon, lat]
                }
            })
            
            # Create geospatial index for new collections
            if lat_val not in processed_lats:
                geo_coll.create_index([('location', GEOSPHERE)])
                # Add to unified index
                lat_index_coll.insert_one({
                    'latitude': lat_val,
                    'collection': collection_name
                })
                processed_lats.add(lat_val)
    
    wb.close()
    client.close()

def query_by_four_corners(top_left, top_right, bottom_left, bottom_right, 
                         mongo_uri='mongodb://localhost:27017/'):
    """
    Query coordinates within bounding box defined by four (lat, lon) points.
    Uses binary search on sorted latitude index for efficiency.
    """
    # Extract all points (each as (lat, lon))
    points = [top_left, top_right, bottom_left, bottom_right]
    print(f"points = {points}")
    
    # Calculate bounding box
    lats = [float(p[0] )for p in points]
    lons = [float(p[1]) for p in points]
    lat_min, lat_max = min(lats), max(lats)
    lon_min, lon_max = min(lons), max(lons)
    print(f"type lat_min {type(lat_min)} == {lat_min}")
    client = MongoClient(mongo_uri)
    db_raw = client['raw_coordinates_db']
    db_geo = client['geocoordinates_db']
    index_db = client['coordinate_index_db']
    
    # Get sorted latitude index
    lat_index = index_db['latitude_index']
    print(f"lat_index = {lat_index}")
    sorted_lats = sorted([doc['latitude'] for doc in lat_index.find({})])
    print(f"sorted lats = {sorted_lats}")

    
    # Find latitude range using binary search
    left_idx = bisect.bisect_left(sorted_lats, lat_min)
    print(f"left idx(lat_min) = {left_idx}")
    right_idx = bisect.bisect_right(sorted_lats, lat_max)
    print(f"right idx (lat_max))= {right_idx}")

    target_lats = sorted_lats[left_idx:right_idx]
    print(f"target_lats= {target_lats}")
    
    # Get collection names
    collections = [doc['collection'] for doc in 
                  lat_index.find({'latitude': {'$in': target_lats}})]
    
    # Query databases
    raw_results = []
    geo_results = []
    
    for coll_name in collections:
        # Raw database query
        raw_coll = db_raw[coll_name]
        raw_docs = list(raw_coll.find({
            'longitude': {'$gte': lon_min, '$lte': lon_max}
        }))
        for doc in raw_docs:
            doc.pop('_id', None)
        raw_results.extend(raw_docs)
        
        # GeoJSON database query
        geo_coll = db_geo[coll_name]
        geo_docs = list(geo_coll.find({
            'location': {
                '$geoWithin': {
                    '$geometry': {
                        'type': 'Polygon',
                        'coordinates': [[
                            [lon_min, lat_min],
                            [lon_max, lat_min],
                            [lon_max, lat_max],
                            [lon_min, lat_max],
                            [lon_min, lat_min]
                        ]]
                    }
                }
            }
        }))
        for doc in geo_docs:
            doc.pop('_id', None)
        geo_results.extend(geo_docs)
    
    client.close()
    
    return {
        'raw_coordinates': raw_results,
        'geo_coordinates': geo_results
    }


def create_database_datacube(file_path):
    """
    Reads coordinates from Excel, stores in MongoDB with correct (lat, lon) order,
    and creates optimized index collections.
    """
    payload={
                    "db_name": "inscribingLocationsRaw",
                    "collections":  [
                        {
                            "name": "latitude_index",
                            "fields": [{"name":"latitude","type":"number"}, {"name":"collection","type":"string"}]
                        }
                    ]
                }
    # payload = json.dumps(payload)
    print(f"payload dumped ={payload} and type = {type(payload)}")    
    response =  requests.post(create_url,json=payload,headers=headers)
    print(f"status code = {response.status_code}")
    print(f"status content = {response.text}")
    # Track processed latitudes
    
    processed_lats = set()
    count=0
    
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    
    # for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
    #     for col_idx, cell in enumerate(row[1:], start=1):
    #         value = cell.value
    #         if not value or not isinstance(value, str):
    #             continue
                
    #         match = re.match(r'\(([^,]+),\s*([^)]+)\)', value.strip())
    #         if not match:
    #             continue
                
    #         # Extract and swap values: first is latitude, second is longitude
    #         lat_str, lon_str = match.groups()
    #         try:
    #             lat_val = float(lat_str.strip())
    #             lon_val = float(lon_str.strip())
    #         except ValueError:
    #             continue
    #         # Create collection name from latitude
    #         collection_name = f"lat_{lat_val:.6f}"
    #         # field_names = ["latitude","longitude"]
    #         field_names=[{"name":"latitude","type":"number"}, {"name":"longitude","type":"number"}]
    #         temp_collection = {
    #                         "name": collection_name,
    #                         "fields": field_names
    #                     }
    #         payload['collections'].append(temp_collection)
        
    #     response =  requests.post(create_url,data=payload)
    #     print(f"status code = {response.status_code}")
    #     print(f"status content = {response.text}")
    #     payload['collections']=[]
    wb.close()
def add_collections(file_path, database_id):
    payload={
                    
    "database_id": database_id,
    # "db_name":"inscribinglocationsraw",
    "collections": []
                }
    # payload = json.dumps(payload)
    # print(f"payload dumped ={payload} and type = {type(payload)}")    
    # response =  requests.post(create_url,json=payload)
    # print(f"status code = {response.status_code}")
    # print(f"status content = {response.text}")
    # Track processed latitudes
    processed_lats_lst =  [
  "lat_0-899322",
  "lat_0-890328",
  "lat_0-881335",
  "lat_0-872342",
  "lat_0-863349",
  "lat_0-854356",
  "lat_0-845362",
  "lat_0-836369",
  "lat_0-827376",
  "lat_0-818383",
  "lat_0-809389",
  "lat_0-800396",
  "lat_0-791403",
  "lat_0-782410",
  "lat_0-773417",
  "lat_0-764423",
  "lat_0-755430",
  "lat_0-746437",
  "lat_0-737444",
  "lat_0-728451",
  "lat_0-719457",
  "lat_0-710464",
  "lat_0-701471",
  "lat_0-692478",
  "lat_0-683484",
  "lat_0-674491",
  "lat_0-665498",
  "lat_0-656505",
  "lat_0-647512",
  "lat_0-638518",
  "lat_0-629525",
  "lat_0-620532",
  "lat_0-611539",
  "lat_0-602545",
  "lat_0-593552",
  "lat_0-584559",
  "lat_0-575566",
  "lat_0-566573",
  "lat_0-557579",
  "lat_0-548586",
  "lat_0-539593",
  "lat_0-530600",
  "lat_0-521607",
  "lat_0-512613",
  "lat_0-503620",
  "lat_0-494627",
  "lat_0-485634",
  "lat_0-476640",
  "lat_0-467647",
  "lat_0-458654",
  "lat_0-449661",
  "lat_0-440668",
  "lat_0-431674",
  "lat_0-422681",
  "lat_0-413688",
  "lat_0-404695",
  "lat_0-395702",
  "lat_0-386708"
]
    processed_lats = set(processed_lats_lst)
    count=0
    
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
        for col_idx, cell in enumerate(row[1:], start=1):
            value = cell.value
            if not value or not isinstance(value, str):
                continue
                
            match = re.match(r'\(([^,]+),\s*([^)]+)\)', value.strip())
            if not match:
                continue
                
            # Extract and swap values: first is latitude, second is longitude
            lat_str, lon_str = match.groups()
            try:
                lat_val = float(lat_str.strip())
                lon_val = float(lon_str.strip())
            except ValueError:
                continue
            # Create collection name from latitude
            collection_name = f"lat_{lat_val:.6f}"
            collection_name=collection_name.replace('.', '-')
            if collection_name in processed_lats:
                continue
            # field_names = ["latitude","longitude"]
            field_names=[{"name":"latitude","type":"number"}, {"name":"longitude","type":"number"}]
            temp_collection = {
                            "name": collection_name,
                            "fields": field_names
                        }
            payload['collections'].append(temp_collection)
            processed_lats.add(collection_name)
        print(f"Lenght of collections now =====> {len(payload['collections'])} <=======")
        # print(f"collections = {payload['collections']} about to be inserted")
        if len(payload["collections"]):
            print(f"Lenght of collections = {len(payload['collections'])} about to be inserted") 
            response =  requests.post(add_colls_url,json=payload, headers=headers)
            count +=1
            print(f"status code of count {count} = {response.status_code}")
            print(f"status content of count {count}= {response.text}")
            payload['collections']=[]
            print(f"Lenght of collections = {len(payload['collections'])} after insert")
    wb.close()

def insert_data_datacube(file_path, database_id):
    """
    Reads coordinates from Excel, stores in MongoDB with correct (lat, lon) order,
    and creates optimized index collections.
    """    
    
    # Create unified index collection
    # index_db = client['coordinate_index_db']
    # lat_index_coll = index_db['latitude_index']
    # lat_index_coll.create_index("latitude")
    lat_index_payload = {
    "database_id": database_id,
    "collection_name": "latitude_index",
    "data": []
    }
    
    
    # Track processed latitudes
   
    processed_lats = set()
    
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb.active
    count=0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2)):
        for col_idx, cell in enumerate(row[1:], start=1):
            value = cell.value
            if not value or not isinstance(value, str):
                continue
                
            match = re.match(r'\(([^,]+),\s*([^)]+)\)', value.strip())
            if not match:
                continue
                
            # Extract and swap values: first is latitude, second is longitude
            lat_str, lon_str = match.groups()
            try:
                lat_val = float(lat_str.strip())
                lon_val = float(lon_str.strip())
            except ValueError:
                continue
                
            # Create collection name from latitude
            collection_name = f"lat_{lat_val:.6f}"
            collection_name=collection_name.replace('.', '-')
            
            # Insert into raw database
            raw_payload = {
            "database_id": database_id,
            "collection_name": collection_name,
            "data": [
                {
                'latitude': lat_val,
                'longitude': lon_val
            }
            ]
            }
            response = requests.post(crud_url,json=raw_payload, headers=headers)
            count +=1
            print(f"raw payload count ={count}  status code = {response.status_code}")
            print(f"raw payload count ={count} status content = {response.text}")

            
            
            # Create geospatial index for new collections
            if lat_val not in processed_lats:
                # geo_coll.create_index([('location', GEOSPHERE)])
                # # Add to unified index
                # lat_index_coll.insert_one({
                #     'latitude': lat_val,
                #     'collection': collection_name
                # })
                lat_index_payload["data"].append({
                    'latitude': lat_val,
                    'collection': collection_name
                })
                processed_lats.add(lat_val)
    res=requests.post(crud_url, json=lat_index_payload, headers=headers)
    print(f"lat index res status code = {res.status_code}")
    print(f"lat index status content = {res.text}")
    wb.close()

def query_by_four_corners_datacube(top_left, top_right, bottom_left, bottom_right, 
                         database_id):
    """
    Query coordinates within bounding box defined by four (lat, lon) points.
    Uses binary search on sorted latitude index for efficiency.
    """
    # Extract all points (each as (lat, lon))
    points = [top_left, top_right, bottom_left, bottom_right]
    print(f"points = {points}")
    
    # Calculate bounding box
    lats = [float(p[0] )for p in points]
    lons = [float(p[1]) for p in points]
    lat_min, lat_max = min(lats), max(lats)
    lon_min, lon_max = min(lons), max(lons)
    print(f"type lat_min {type(lat_min)} == {lat_min}")
    
    index_parameters = f"?database_id={database_id}&collection_name=latitude_index&filters={{}}&page=1&page_size=200"
    url=crud_url+index_parameters
    # index_parameters= /api/crud/ ? database_id=507f1f77bcf86cd799439011 & collection_name=users & filters={"age": {"$gt": 30}} & page=1 & page_size=50
    # client = MongoClient(mongo_uri)
    # db_raw = client['raw_coordinates_db']
    # db_geo = client['geocoordinates_db']
    # index_db = client['coordinate_index_db']
    
    # Get sorted latitude index
    # lat_index = index_db['latitude_index']
    # print(f"lat_index = {lat_index}")
    lat_index = requests.get(url)
    sorted_lats = sorted([doc['latitude'] for doc in lat_index['data']])
    print(f"sorted lats = {sorted_lats}")
    # Find latitude range using binary search
    left_idx = bisect.bisect_left(sorted_lats, lat_min)
    print(f"left idx(lat_min) = {left_idx}")
    right_idx = bisect.bisect_right(sorted_lats, lat_max)
    print(f"right idx (lat_max))= {right_idx}")
    target_lats = sorted_lats[left_idx:right_idx]
    print(f"target_lats= {target_lats}")
    # Get collection names
    collections = [doc['collection'] for doc in 
                  lat_index.find({'latitude': {'$in': target_lats}})]
    # Query databases
    raw_results = []
    for coll_name in collections:
        # Raw database query
        # raw_coll = db_raw[coll_name]
        fil={'longitude': {'$gte': lon_min, '$lte': lon_max}}
        raw_parameters = f"?database_id={database_id}&collection_name={coll_name}&filters={fil}&page=1&page_size=200"
        
        raw_url= crud_url+raw_parameters
        raw_coll= requests.get(raw_url)
        raw_docs = list(raw_coll['data'].find({
            'longitude': {'$gte': lon_min, '$lte': lon_max}
        }))
        for doc in raw_docs:
            doc.pop('_id', None)
        raw_results.extend(raw_docs)
        
        # GeoJSON database query
        # geo_coll = db_geo[coll_name]
        # geo_docs = list(geo_coll.find({
        #     'location': {
        #         '$geoWithin': {
        #             '$geometry': {
        #                 'type': 'Polygon',
        #                 'coordinates': [[
        #                     [lon_min, lat_min],
        #                     [lon_max, lat_min],
        #                     [lon_max, lat_max],
        #                     [lon_min, lat_max],
        #                     [lon_min, lat_min]
        #                 ]]
        #             }
        #         }
        #     }
        # }))
        # for doc in geo_docs:
        #     doc.pop('_id', None)
        # geo_results.extend(geo_docs)
        
    return {
        'raw_coordinates': raw_results,
    }
database_detail =  {"success":True,"database":{"id":"68b16e0f77a6f0d3e0dd3770","name":"inscribinglocationsraw"},
                    "collections":[{"name":"latitude_index","created":True,"exists":False,"error":None}]}
culprit_filename="scaled_coordinates_200_200_0_0.xlsx"
culprit_database_id= config("DATABASE_ID")

# create_database_datacube(culprit_filename)
# add_collections(culprit_filename,culprit_database_id)
insert_data_datacube(culprit_filename,culprit_database_id)
# print(f"Length == {type(colls)}")
# print(f"Length == {colls.keys()}")
# print(f"Length == {len(colls.json()['collections'])}")
# print(f"Length == {colls.json().keys()}")
# store_coordinates(culprit_filename)
# top_left=(13.133809921183700,77.645992221183800)
# top_right=(13.133809921183700,77.286263578816300)
# bottom_left=(12.774081278816300,77.286263578816300)
# bottom_right=(12.774081278816300,77.645992221183800)
# print(query_by_four_corners(top_left, top_right, bottom_left, bottom_right))